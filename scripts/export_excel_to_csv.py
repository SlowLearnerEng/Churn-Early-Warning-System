#!/usr/bin/env python3
"""
Export all sheets from indiamart_churn_dataset_v2.xlsx to data/csv/ as CSVs.

Usage:
    python scripts/export_excel_to_csv.py
"""
import csv
import openpyxl
from pathlib import Path

EXCEL_PATH = Path("indiamart_churn_dataset_v2.xlsx")
OUT_DIR    = Path("data/csv")
OUT_DIR.mkdir(parents=True, exist_ok=True)

def fmt(val):
    if val is None:
        return ""
    return str(val)

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    if not rows:
        continue
    out_path = OUT_DIR / f"dataset - {sheet_name}.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in rows:
            w.writerow([fmt(v) for v in row])
    print(f"  {sheet_name:35s} -> {out_path.name}  ({len(rows)-1} rows)")

print(f"\nAll {len(wb.sheetnames)} sheets exported to {OUT_DIR}/")
